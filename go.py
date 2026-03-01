import os
import sys
import json
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from datetime import datetime
from models.store import Store
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from PIL import Image

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class Go_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str, chrome_path: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.browser = webdriver.Chrome(service=ChromeService(chrome_path), options=self.chrome_options)
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            auth_token: str = ''
            collections: list[dict] = list()

            self.browser.get(store.link)
            self.wait_until_browsing()

            if self.login(store.username, store.password):
                if self.wait_until_element_found(20, 'xpath', '//ul[@data-submenu-title="Collections"]'):

                    if not collections: collections = self.get_collections()
                    
                    
                    for brand_with_type in brands_with_types:
                        brand: str = brand_with_type['brand']
                        brand_code: str = str(brand_with_type['code']).strip()
                        print(f'Brand: {brand}')

                        brand_collection = self.get_brand_collection(collections, brand, brand_code)
                        

                        for glasses_type in brand_with_type['glasses_type']:
                            start_time = datetime.now()
                            auth_token = self.get_token()
                            brand_data = self.get_brand_data(brand_collection, glasses_type)

                            total_products = len(brand_data) if brand_data else 0
                            print(f'Type: {glasses_type} | Total products: {total_products}')
                            print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                

                            if brand_data:
                                self.printProgressBar(0, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                for scraped_products, product_data in enumerate(brand_data):

                                    self.printProgressBar(scraped_products + 1, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                    product_json = self.get_product_data(product_data, brand_collection, auth_token)
                                    if product_json:
                                        self.normalize_product_data(brand, glasses_type, product_data, product_json)
                                        self.save_to_json(self.data)
                                    

                            self.save_to_json(self.data)
                            end_time = datetime.now()

                            print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                            print('Duration: {}\n'.format(end_time - start_time))

                            self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                            self.print_logs('Duration: {}\n'.format(end_time - start_time))

            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in Go_Scraper controller: {e}')
            self.print_logs(f'Exception in Go_Scraper controller: {e}')
        finally: 
            self.browser.quit()
            self.save_to_json(self.data)

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def select_language(self) -> None:
        try:
            language_select = Select(self.browser.find_element(By.XPATH, '//select[@id="language"]'))
            language_select.select_by_value('en_US')
        except Exception as e:
            if self.DEBUG: print(f'Exception in select_language: {e}')
            self.print_logs(f'Exception in select_language: {e}')

    def login(self, username: str, password: str) -> bool:
        login_flag = False
        try:
            self.select_language()
            if self.wait_until_element_found(50, 'xpath', '//input[@name="username"]'):
                # try:
                #     button = WebDriverWait(self.browser, 10).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Accept"]')))
                #     button.click()
                # except: pass

                self.browser.find_element(By.XPATH, '//input[@name="username"]').send_keys(username)
                self.browser.find_element(By.XPATH, '//input[@name="password"]').send_keys(password)
                try:
                    button = WebDriverWait(self.browser, 50).until(EC.element_to_be_clickable((By.XPATH, '//button[@type="submit"]')))
                    button.click()

                    WebDriverWait(self.browser, 50).until(EC.presence_of_element_located((By.XPATH, '//ul[@data-submenu-title="Collections"]')))
                    login_flag = True
                except Exception as e:
                    self.print_logs(str(e))
                    if self.DEBUG: print(str(e))
                    else: pass
        except Exception as e:
            self.print_logs(f'Exception in login: {str(e)}')
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            else: pass
        finally: return login_flag

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def get_collections(self) -> list[dict]:
        collections = list()
        try:
            for _ in range(1, 10):
                value = self.browser.execute_script("return window.localStorage.getItem('collections');")
                if value:
                    collections = json.loads(value)
                    break
                sleep(1)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_collections: {e}')
            self.print_logs(f'Exception in get_collections: {e}')
        finally: return collections

    def get_brand_collection(self, collections: list[dict], brand_name: str, brand_code: str) -> dict:
        brand_collection = dict()
        try:
            for collection in collections:
                if str(collection['collection']).strip().lower() == str(brand_name).strip().lower() and str(collection['prefix']).strip().lower() == str(brand_code).strip().lower():
                    brand_collection = collection
                    break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_collection: {e}')
            self.print_logs(f'Exception in get_brand_collection: {e}')
        finally: return brand_collection

    def get_brand_data(self, brand_collection: dict, glasses_type: str) -> list[dict]:
        brand_data: list[dict] = list()
        try:
            headers = {
                'sec-ch-ua-platform': '"Windows"',
                'Referer': 'https://b2b.goeyeweargroup.com/',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
                'Accept': 'application/json, text/plain, */*',
                'sec-ch-ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144"',
                'Content-Type': 'application/json;charset=UTF-8',
                'sec-ch-ua-mobile': '?0',
            }

            json_data = {
                'collection_id': str(brand_collection.get('code')),
                'country': 'IT',
                'gender': [],
                'type': '1' if glasses_type == 'Sunglasses' else '0',
                'sizes': [],
                'shapes': [],
                'material': [],
            }

            response = requests.post('https://b2b-backend.goeyeweargroup.com/api/v1/search', headers=headers, json=json_data)
            if response.status_code == 200:
                brand_data = response.json().get('data')
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_data: {e}')
            self.print_logs((f'Exception in get_brand_data: {e}'))
        finally: return brand_data

    def get_token(self)-> str:
        token = ''
        for _ in range(0, 10):
            try:
                token = self.browser.execute_script( "return window.sessionStorage.getItem('token');")
                if token:
                    return token
                sleep(1)
            except Exception as e:
                if self.DEBUG: print(f'Exception in get_token: {e}')
                self.print_logs((f'Exception in get_token: {e}'))
                sleep(1)
        return token

    def get_product_data(self, product_data: str, brand_collection: dict, auth_token: str) -> dict:
        product_json = dict()
        try:
            headers = {
                'sec-ch-ua-platform': '"Windows"',
                'Authorization': f'Bearer {auth_token}',
                'Referer': 'https://b2b.goeyeweargroup.com/',
                'sec-ch-ua': '"Not(A:Brand";v="8", "Chromium";v="144", "Google Chrome";v="144"',
                'sec-ch-ua-mobile': '?0',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36',
                'Accept': 'application/json, text/plain, */*',
                'Content-Type': 'application/json;charset=UTF-8',
            }

            json_data = {
                'country': 'IT',
                'collection_id': str(brand_collection.get('code')),
                'store_id': 6792,
                'model_name': str(product_data.get('name')).strip().lower(),
                'type': 'current',
            }

            response = requests.post('https://b2b-backend.goeyeweargroup.com/api/v1/model', headers=headers, json=json_data)
            if response.status_code == 200:
                product_json = response.json()

        except Exception as e:
            if self.DEBUG: print(f'Exception in get_product_data: {e}')
            self.print_logs((f'Exception in get_product_data: {e}'))
        finally: return product_json

    def normalize_product_data(self, brand_name: str, glasses_type: str, product_data: dict, product_json: list) -> dict:
        try:

            url = f'https://b2b.goeyeweargroup.com/product/{str(product_data.get("name")).strip().lower()}'
            clip_on = 'Yes' if product_data.get('clipOn') == 1 else 'No'

            for product_child in product_json[0].get(product_data.get("name")).get('childs', []):
                product = Product()
                product.url = url
                product.brand = brand_name
                product.type = glasses_type
                product.frame_code = product_child.get('model')
                product.lens_code = product_child.get('color')

                metafields = Metafields()
                metafields.clip_on = clip_on
                metafields.img_url = product_child.get('images', {}).get('original').replace('\/', '\\')

                varaint = Variant()
                varaint.sku = product_child.get('name')
                varaint.inventory_quantity = 1 if product_child.get('stock') == 'green' else 0
                varaint.wholesale_price = product_child.get('price')
                varaint.size = product_child.get('size')

                product.variants.append(varaint)
                product.metafields = metafields

                self.data.append(product)

        except Exception as e:
            if self.DEBUG: print(f'Exception in normalize_product_data: {e}')
            self.print_logs((f'Exception in normalize_product_data: {e}'))
   
    def save_to_json(self, products: list[Product]) -> None:
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'listing_price': variant.listing_price, 
                        'wholesale_price': variant.wholesale_price,
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                        'weight': variant.weight
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.metafields.img_url },
                        { 'key': 'fitting_info', 'value': product.metafields.fitting_info },
                        { 'key': 'img_360_urls', 'value': product.metafields.img_360_urls },
                        {  'key': 'clip_on', 'value': product.metafields.clip_on }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')
    
    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()


def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                clip_on = ''
                brand = json_d['brand']
                number = str(json_d['number']).strip().upper()
                if '/' in number: number = number.replace('/', '-').strip()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                lens_color = str(json_d['lens_color']).strip().title()
                glasses_type = str(json_d['type']).strip()
                
                for json_metafiels in json_d['metafields']:
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                    if json_metafiels['key'] == 'clip_on': clip_on = str(json_metafiels['value']).strip()
                for json_variant in json_d['variants']:
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    image_filename = f'Images/{sku}.jpg'
                    if not os.path.exists(image_filename): 
                        image_attachment = download_image(img_url)
                        if image_attachment:
                            with open(image_filename, 'wb') as f: f.write(image_attachment)
                            crop_downloaded_image(f'Images/{sku}.jpg')
                    data.append([frame_code, lens_color, brand, glasses_type, sku, wholesale_price, clip_on])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1680
        new_height = 1020
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Model Code')
    worksheet.cell(row=1, column=2, value='Lens Code')
    worksheet.cell(row=1, column=3, value='Brand')
    worksheet.cell(row=1, column=4, value="Glasses Type")
    worksheet.cell(row=1, column=5, value='SKU')
    worksheet.cell(row=1, column=6, value='Wholesale Price')
    worksheet.cell(row=1, column=7, value="Clip On")
    worksheet.cell(row=1, column=8, value="Image")


    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])

        image = f'Images/{d[-3]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='H'+str(new_index))

    workbook.save('Go Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    # download chromedriver.exe with same version and get its path
    if os.path.exists('Go Results.xlsx'): os.remove('Go Results.xlsx')

    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Go start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/Go.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Go Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'

    chrome_path = ''
    if not chrome_path:
        chrome_path = ChromeDriverManager().install()
        if 'chromedriver.exe' not in chrome_path:
            chrome_path = str(chrome_path).split('/')[0].strip()
            chrome_path = f'{chrome_path}\\chromedriver.exe'
    
    Go_Scraper(DEBUG, result_filename, logs_filename, chrome_path).controller(store, brands)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    print('Downloading images....')
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)
    
    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
